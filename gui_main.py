import base64
import hashlib
import os
import traceback

import pandas as pd
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from openpyxl.styles import Font, Side, Border, Alignment, PatternFill
from openpyxl.utils import get_column_letter as n2a

from gui_config import QTabWidget, cur_path
from gui_data_check import is_must_check, check_all
from gui_subtable import tab_data, buttonLayoutGen, tabLayoutGenerate, tabLayoutGenerate_2, tabLayoutGenerate_3
from gui_subtable_multi import tabLayoutGenerate_4, tabLayoutGenerate_5
from logger import logger


class UiForm(QtWidgets.QWidget):

    def __init__(self, parent, w, h):
        super(UiForm, self).__init__(parent)
        self.main_form = None
        self.height = None
        self.width = None
        self.w = w
        self.h = h
        self.btn_reset = None
        self.btn_submit = None  # 提交按钮
        self.text_label_dict = {}  # 保存内容2的组件
        self.buttonWidget = None
        self.btn_next = None  # 下一页
        self.btn_prov = None  # 上一页
        self.textWidget = None
        self.data_table = tab_data.data  # 配置数据的数据类
        self.tab_name_list = self.data_table.tab_name
        self.gridLayout = None  # 窗口的主要布局
        self.tabWidget = None  # tab分页的组件
        self.tab_list = []  # 保存的tab分页内容
        self.tab_gen_func = [tabLayoutGenerate, tabLayoutGenerate_2, tabLayoutGenerate_3, tabLayoutGenerate_4,
                             tabLayoutGenerate_5, tabLayoutGenerate]

    def eventFilter(self, obj, event):
        if event.type() == QtCore.QEvent.Resize:
            self.width = self.main_form.width()
            self.height = self.main_form.height()
            return super().eventFilter(obj, event)

    def setupUi(self, main_form):
        main_form.setObjectName("Form")
        main_form.resize(860, 630)
        self.main_form = main_form
        self.width = main_form.width()
        self.height = main_form.height()
        # main_form.setFixedSize(main_form.width(), main_form.height())
        # 设置全局布局
        self.gridLayout = QtWidgets.QGridLayout()
        main_form.setLayout(self.gridLayout)
        # Form.installEventFilter(mousePressEvent)

        self.gridLayout.setObjectName("gridLayout")
        # 添加tab Widget
        self.tabWidget = QTabWidget()
        self.tabWidget.in_signal.connect(lambda: self.tabWidget.setFocus())
        self.gridLayout.addWidget(self.tabWidget)


        # self.tabWidget.resize(850, 770)
        self.tabWidget.setTabPosition(QtWidgets.QTabWidget.North)
        self.tabWidget.setTabShape(QtWidgets.QTabWidget.Rounded)
        self.tabWidget.setObjectName("tabWidget")

        # 添加提示信息
        textWidget = QtWidgets.QTextBrowser()
        textWidget.setFixedHeight(self.width//5)
        self.textWidget = textWidget
        self.gridLayout.addWidget(self.textWidget)

        # 添加按钮
        buttonWidget = QtWidgets.QWidget()
        btn_layout, btn_list = buttonLayoutGen()
        self.btn_prov = btn_list[0]
        self.btn_next = btn_list[1]
        self.btn_submit = btn_list[2]
        self.btn_reset = btn_list[3]
        self.btn_prov.clicked.connect(self.tab_prov)
        self.btn_next.clicked.connect(self.tab_next)
        self.btn_submit.clicked.connect(self.tab_submit)
        self.btn_reset.clicked.connect(lambda: self.tab_reset(main_form))
        buttonWidget.setLayout(btn_layout)
        self.buttonWidget = buttonWidget
        self.gridLayout.addWidget(self.buttonWidget)

        for index, i in enumerate(self.tab_name_list):
            tab_layout, text_widget_list, text_label_dict = self.tab_gen_func[index](self, i)
            # tab_layout, text_widget_list, text_label_dict = tabLayoutGenerate(i)
            # 内容2的组件
            self.text_label_dict.update({i: text_label_dict})
            # 绑定信号到保存函数，焦点改变时刷新数据
            [wd.in_signal.connect(
                lambda in_dict: is_must_check(tab_data, in_dict, self.text_label_dict)) if wd is not None else wd for wd
             in text_widget_list]
            tab = QtWidgets.QWidget()
            tab.setObjectName(i)
            tab.setLayout(tab_layout)
            self.tabWidget.addTab(tab, i)
            self.tab_list.append(tab)

            # ly.addLayout()
        # 改变tab时重新设置焦点，避免lineEdit焦点一直存在
        self.tabWidget.currentChanged.connect(lambda: self.tabWidget.setFocus())

        # self.gridLayout.addWidget(self.tabWidget, 0, 0, 0, 0)
        # self.gridLayout.addLayout(tab_layout, 0, 1, 0, 1)
        try:
            self.retranslateUi(main_form)
            # self.tabWidget.setCurrentIndex(0)
            QtCore.QMetaObject.connectSlotsByName(main_form)
        except Exception as e:
            logger.error(e)
            logger.error(traceback.format_exc())
            traceback.print_exc(file=open('err_log.log', 'w+'))
            # traceback.format_exception

    def tab_prov(self):
        tab_index = self.tabWidget.currentIndex()
        if tab_index == 0:
            return
        self.tabWidget.setCurrentIndex(tab_index - 1)

    def tab_next(self):
        tab_index = self.tabWidget.currentIndex()
        if tab_index > len(self.tab_name_list):
            return
        self.tabWidget.setCurrentIndex(tab_index + 1)

    def tab_submit(self):
        check_pass_flag, check_msg, df_data = check_all(tab_data)
        self.textWidget.setText(check_msg)
        if check_pass_flag:
            project_name = tab_data.data.get('rec_info').get('项目基础信息').get('项目名称')[0]
            file_path = os.path.join(cur_path, 'ICT项目方案预审表-{}.xlsx'.format(project_name))
            self.textWidget.setText('校验通过\r\n文件输出路径：{}'.format(file_path))
            try:
                writer = pd.ExcelWriter(file_path)
                # max_row = ws.max_row
                res_data = pd.DataFrame(data=df_data, columns=['序号', '模块', '类型', '内容1', '内容2']).set_index(
                    ['序号', '模块', '类型', '内容1'])
                col_width = [5.0, 17.75, 38.5, 56.875, 35.0]

                res_data.to_excel(writer, startrow=1)
                ws = writer.sheets['Sheet1']
                max_col = ws.max_column
                merge_cell = '{}1:{}1'.format(n2a(1), n2a(max_col))
                ws.merge_cells(merge_cell)
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws['A1'].fill = PatternFill(fill_type='solid', fgColor="FFFF00")
                ws['A1'].border = Border(left=Side(style='thin', color='FF000000'),
                                         right=Side(style='thin', color='FF000000'),
                                         top=Side(style='thin', color='FF000000'),
                                         bottom=Side(style='thin', color='FF000000'),
                                         diagonal=Side(style='thin', color='FF000000'),
                                         diagonal_direction=0,
                                         outline=Side(style='thin', color='FF000000'),
                                         vertical=Side(style='thin', color='FF000000'),
                                         horizontal=Side(style='thin', color='FF000000'))
                ws['A1'].value = 'ICT项目方案预审表'
                ws['A1'].font = Font(bold=True)
                for index_r, row_data in enumerate(ws.rows):
                    for index, col_cell in enumerate(row_data):
                        if index_r == 0:
                            ws.column_dimensions[n2a(index + 1)].width = col_width[index]
                            ws[f'{n2a(index + 1)}2'].fill = PatternFill(fill_type='solid', fgColor="FFFF00")
                        col_cell.font = Font(bold=False)
                        col_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                        #         print(col_cell.border)
                        #         break
                        if index > 2 and index_r > 0:
                            col_cell.border = Border(left=Side(style='thin', color='FF000000'),
                                                     right=Side(style='thin', color='FF000000'),
                                                     top=Side(style='thin', color='FF000000'),
                                                     bottom=Side(style='thin', color='FF000000'),
                                                     diagonal=Side(style='thin', color='FF000000'),
                                                     diagonal_direction=0,
                                                     outline=Side(style='thin', color='FF000000'),
                                                     vertical=Side(style='thin', color='FF000000'),
                                                     horizontal=Side(style='thin', color='FF000000'))
                writer.save()
            except Exception as e:
                if isinstance(e, PermissionError):
                    self.textWidget.append('导出失败\r\n无访问权限，请检查Excel文件是否关闭后再次点击导出')
                self.textWidget.append('错误日志详见err_log.log')
                logger.error(e)
                logger.error(traceback.format_exc())
                traceback.print_exc(file=open('err_log.log', 'w+'))
        # tab_data.save()
        # pd.ExcelFile

        # print(self.data_table['rec_info'])

    @QtCore.pyqtSlot()
    def tab_reset(self, main_form):
        # self.__init__()
        QtCore.QCoreApplication.exit(11950)
        # self.setupUi(main_form)

    def retranslateUi(self, main_form):
        """
        文本设置
        :param main_form:
        :return:
        """
        _translate = QtCore.QCoreApplication.translate
        main_form.setWindowTitle(_translate("Form", "ICT预审表"))
        for index, tab in enumerate(self.tab_list):
            self.tabWidget.setTabText(self.tabWidget.indexOf(tab),
                                      _translate(str(index), '{}.{}'.format(index + 1, self.tab_name_list[index])))


class pwdUI(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super(pwdUI, self).__init__(parent)

        self.setObjectName('login')
        # self.resize(w, h)
        # self.setFixedSize(self.width(), self.height())

        # 输入密码框
        flo = QtWidgets.QHBoxLayout()
        e1 = QtWidgets.QLineEdit()
        self.pwd = e1
        BtnOk = QtWidgets.QPushButton("确 定")
        e1.setEchoMode(QtWidgets.QLineEdit.Password)  # 设置密码不可见
        BtnOk.clicked.connect(self.textchanged)
        e1.returnPressed.connect(self.textchanged)
        flo.addWidget(QtWidgets.QLabel("请输入密码："))
        flo.addWidget(e1)
        flo.addWidget(BtnOk)
        self.setLayout(flo)

    # 核对密码是否正确
    def textchanged(self):
        s = hashlib.sha256()  # Get the hash algorithm.
        s.update(base64.b64encode(self.pwd.text().encode('utf-8')))  # Hash the data.
        b = s.hexdigest()
        if b == "374706e7b276d2ecc4204ef00bd9de68a402e461ce9005d46afc6c4ebd77d12b":
            QtCore.QCoreApplication.exit(1)  # 关闭登陆界面
        else:
            self.pwd.clear()
            QMessageBox.warning(self, "提示", "密码错误", QMessageBox.Yes | QMessageBox.No)
